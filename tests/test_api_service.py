from __future__ import annotations

import json
from io import BytesIO
from pathlib import Path
import sqlite3

from fastapi.testclient import TestClient
from openpyxl import load_workbook

import pedestrian_line_counter.api as api_module
import pedestrian_line_counter.service as service_module
from pedestrian_line_counter.api import MutationAuthConfig, create_app
from pedestrian_line_counter.config import ServiceConfig, SpoolRetentionConfig
from pedestrian_line_counter.event_uploader import RetryConfig, SyncSummary, UploaderConfig
from pedestrian_line_counter.review_store import ReviewStore
from pedestrian_line_counter.ui_auth import UiAuthConfig


def _write_run(
    root: Path,
    *,
    day: str,
    run_uid: str,
    started_at_utc: str,
    occurred_at_utc: str,
    camera_id: str = "cam_01",
    occurred_at_local: str | None = None,
    direction: str = "A_TO_B",
    completed_at_utc: str | None = None,
    last_error: str | None = None,
    last_error_at_utc: str | None = None,
    in_progress_last_sync_at_utc: str | None = None,
    lifecycle_status: str = "stopped",
    class_name: str = "truck",
    run_class_names: dict[int, str] | None = None,
    thumb_relpath: str | None = "thumbs/e1.jpg",
    scene_relpath: str | None = "scene/e1.jpg",
) -> Path:
    run_dir = root / day / run_uid
    (run_dir / "thumbs").mkdir(parents=True, exist_ok=True)
    (run_dir / "scene").mkdir(parents=True, exist_ok=True)

    run_json = {
        "run_uid": run_uid,
        "site_id": "site_a",
        "camera_id": camera_id,
        "started_at_utc": started_at_utc,
        "updated_at_utc": started_at_utc,
        "ended_at_utc": started_at_utc,
        "source": {"type": "rtsp", "value": "rtsp://camera"},
        "line_mode": "line",
        "class_names": (
            {str(class_id): name for class_id, name in run_class_names.items()}
            if run_class_names is not None
            else {}
        ),
        "report_csv_relpath": "report.csv",
        "health_summary": {
            "lifecycle_status": lifecycle_status,
            "frames_total": 100,
            "frames_processed": 50,
            "events_emitted_total": 1,
            "count_a_to_b": 1,
            "count_b_to_a": 0,
            "effective_fps": 12.5,
            "processed_fps": 6.25,
        },
    }
    (run_dir / "run.json").write_text(json.dumps(run_json), encoding="utf-8")
    (run_dir / "status.json").write_text(
        json.dumps(
            {
                "run_uid": run_uid,
                "updated_at_utc": occurred_at_utc,
                "health_summary": run_json["health_summary"],
            }
        ),
        encoding="utf-8",
    )
    (run_dir / "events.jsonl").write_text(
        json.dumps(
            {
                "event_uid": f"{run_uid}_e1",
                "run_uid": run_uid,
                "site_id": "site_a",
                "camera_id": camera_id,
                "occurred_at_utc": occurred_at_utc,
                "occurred_at_local": occurred_at_local,
                "frame_index": 90,
                "video_time_s": 3.0,
                "direction": direction,
                "track_id": 1,
                "class_id": 2,
                "class_name": class_name,
                "confidence": 0.91,
                "thumb_relpath": thumb_relpath,
                "scene_relpath": scene_relpath,
            }
        )
        + "\n",
        encoding="utf-8",
    )
    if thumb_relpath:
        thumb_path = run_dir / Path(thumb_relpath)
        thumb_path.parent.mkdir(parents=True, exist_ok=True)
        thumb_path.write_bytes(b"jpg")
    if scene_relpath:
        scene_path = run_dir / Path(scene_relpath)
        scene_path.parent.mkdir(parents=True, exist_ok=True)
        scene_path.write_bytes(b"jpg")

    state = {}
    if completed_at_utc is not None:
        state["completed_at_utc"] = completed_at_utc
    if last_error is not None:
        state["last_error"] = last_error
    if last_error_at_utc is not None:
        state["last_error_at_utc"] = last_error_at_utc
    if in_progress_last_sync_at_utc is not None:
        state["in_progress_last_sync_at_utc"] = in_progress_last_sync_at_utc
    if state:
        state["run_uid"] = run_uid
        (run_dir / ".portal_upload_state.json").write_text(json.dumps(state), encoding="utf-8")

    return run_dir


def _login_ui(client: TestClient, *, username: str = "admin", password: str = "secret") -> None:
    resp = client.post(
        "/api/auth/login",
        json={
            "username": username,
            "password": password,
        },
    )
    assert resp.status_code == 200
    assert resp.json()["ok"] is True


def test_healthz_and_status_expose_spool_state(tmp_path) -> None:
    _write_run(
        tmp_path,
        day="2026-03-11",
        run_uid="run_a",
        started_at_utc="2026-03-11T10:00:00Z",
        occurred_at_utc="2026-03-11T10:05:00Z",
        completed_at_utc="2026-03-11T10:06:00Z",
    )
    _write_run(
        tmp_path,
        day="2026-03-10",
        run_uid="run_b",
        started_at_utc="2026-03-10T09:00:00Z",
        occurred_at_utc="2026-03-10T09:05:00Z",
        last_error="temporary failure",
    )

    client = TestClient(create_app(spool_dir=tmp_path))

    health = client.get("/healthz")
    assert health.status_code == 200
    assert health.json()["ok"] is True
    assert health.json()["spool_exists"] is True

    status = client.get("/status")
    assert status.status_code == 200
    payload = status.json()
    assert payload["runs_total"] == 2
    assert payload["delivery_state_counts"]["completed"] == 1
    assert payload["delivery_state_counts"]["failed"] == 1
    assert payload["sync_overview"]["status_cards"][2]["label"] == "Gagal"
    assert payload["sync_overview"]["status_cards"][2]["value"] == 1
    assert payload["latest_run"]["run_uid"] == "run_a"
    assert payload["latest_run"]["updated_at_utc"] == "2026-03-11T10:05:00Z"


def test_operational_read_endpoints_require_ui_auth_when_enabled(tmp_path) -> None:
    _write_run(
        tmp_path,
        day="2026-03-11",
        run_uid="run_secure",
        started_at_utc="2026-03-11T10:00:00Z",
        occurred_at_utc="2026-03-11T10:05:00Z",
        completed_at_utc="2026-03-11T10:06:00Z",
    )

    retention_cfg = SpoolRetentionConfig(
        enabled=True,
        max_age_days=45,
        protect_incomplete_runs=True,
        state_filename=".portal_upload_state.json",
    )
    client = TestClient(
        create_app(
            spool_dir=tmp_path,
            retention_cfg=retention_cfg,
            ui_auth_cfg=UiAuthConfig(username="admin", password="secret"),
        )
    )

    health = client.get("/healthz")
    assert health.status_code == 200

    for path in (
        "/status",
        "/metrics",
        "/config",
        "/runs/recent",
        "/events/recent",
        "/events/run_secure_e1",
        "/retention/preview",
    ):
        resp = client.get(path)
        assert resp.status_code == 401
        assert resp.json()["detail"] == "ui_auth_required"

    _login_ui(client)

    assert client.get("/status").status_code == 200
    assert client.get("/metrics").status_code == 200
    assert client.get("/config").status_code == 200
    assert client.get("/runs/recent").status_code == 200
    assert client.get("/events/recent").status_code == 200
    assert client.get("/events/run_secure_e1").status_code == 200
    assert client.get("/retention/preview").status_code == 200


def test_runs_recent_redacts_rtsp_source_values(tmp_path) -> None:
    _write_run(
        tmp_path,
        day="2026-03-11",
        run_uid="run_secure",
        started_at_utc="2026-03-11T10:00:00Z",
        occurred_at_utc="2026-03-11T10:05:00Z",
        completed_at_utc="2026-03-11T10:06:00Z",
    )
    run_dir = tmp_path / "2026-03-11" / "run_secure"
    run_meta = json.loads((run_dir / "run.json").read_text(encoding="utf-8"))
    run_meta["source"] = {"type": "rtsp", "value": "rtsp://admin:Secret123@192.168.1.50:554/stream1"}
    run_meta["source_type"] = "rtsp"
    run_meta["source_value"] = "rtsp://admin:Secret123@192.168.1.50:554/stream1"
    (run_dir / "run.json").write_text(json.dumps(run_meta), encoding="utf-8")

    client = TestClient(
        create_app(
            spool_dir=tmp_path,
            ui_auth_cfg=UiAuthConfig(username="admin", password="secret"),
        )
    )
    _login_ui(client)

    payload = client.get("/runs/recent").json()

    assert payload["items"][0]["source_type"] == "rtsp"
    assert payload["items"][0]["source_value"] == "camera:cam_01"


def test_event_evidence_routes_require_ui_auth_and_serve_event_images(tmp_path) -> None:
    _write_run(
        tmp_path,
        day="2026-03-11",
        run_uid="run_evidence",
        started_at_utc="2026-03-11T10:00:00Z",
        occurred_at_utc="2026-03-11T10:05:00Z",
    )

    client = TestClient(
        create_app(
            spool_dir=tmp_path,
            ui_auth_cfg=UiAuthConfig(username="admin", password="secret"),
        )
    )

    unauth_thumb = client.get("/events/run_evidence_e1/thumbnail")
    assert unauth_thumb.status_code == 401
    assert unauth_thumb.json()["detail"] == "ui_auth_required"

    unauth_scene = client.get("/events/run_evidence_e1/scene")
    assert unauth_scene.status_code == 401
    assert unauth_scene.json()["detail"] == "ui_auth_required"

    _login_ui(client)

    thumb = client.get("/events/run_evidence_e1/thumbnail")
    assert thumb.status_code == 200
    assert thumb.content == b"jpg"
    assert thumb.headers["content-type"] == "image/jpeg"

    scene = client.get("/events/run_evidence_e1/scene")
    assert scene.status_code == 200
    assert scene.content == b"jpg"
    assert scene.headers["content-type"] == "image/jpeg"


def test_event_evidence_routes_reject_non_evidence_files(tmp_path) -> None:
    run_dir = _write_run(
        tmp_path,
        day="2026-03-11",
        run_uid="run_bad_evidence",
        started_at_utc="2026-03-11T10:00:00Z",
        occurred_at_utc="2026-03-11T10:05:00Z",
        thumb_relpath="report.csv",
        scene_relpath=None,
    )
    (run_dir / "report.csv").write_text("not an image", encoding="utf-8")

    client = TestClient(
        create_app(
            spool_dir=tmp_path,
            ui_auth_cfg=UiAuthConfig(username="admin", password="secret"),
        )
    )
    _login_ui(client)

    blocked = client.get("/events/run_bad_evidence_e1/thumbnail")
    assert blocked.status_code == 404
    assert blocked.json()["detail"] == "evidence_not_found"


def test_recent_runs_and_events_are_sorted_newest_first(tmp_path) -> None:
    _write_run(
        tmp_path,
        day="2026-03-09",
        run_uid="run_old",
        started_at_utc="2026-03-09T09:00:00Z",
        occurred_at_utc="2026-03-09T09:05:00Z",
    )
    _write_run(
        tmp_path,
        day="2026-03-11",
        run_uid="run_new",
        started_at_utc="2026-03-11T09:00:00Z",
        occurred_at_utc="2026-03-11T09:05:00Z",
    )

    client = TestClient(create_app(spool_dir=tmp_path))

    runs = client.get("/runs/recent", params={"limit": 2})
    assert runs.status_code == 200
    run_items = runs.json()["items"]
    assert [item["run_uid"] for item in run_items] == ["run_new", "run_old"]

    events = client.get("/events/recent", params={"limit": 2})
    assert events.status_code == 200
    event_items = events.json()["items"]
    assert [item["run_uid"] for item in event_items] == ["run_new", "run_old"]
    assert event_items[0]["thumb_path"].endswith("thumbs/e1.jpg")


def test_metrics_and_config_expose_runtime_configuration(tmp_path) -> None:
    _write_run(
        tmp_path,
        day="2026-03-09",
        run_uid="run_old",
        started_at_utc="2026-03-09T09:00:00Z",
        occurred_at_utc="2026-03-09T09:05:00Z",
    )
    _write_run(
        tmp_path,
        day="2026-03-11",
        run_uid="run_new",
        started_at_utc="2026-03-11T09:00:00Z",
        occurred_at_utc="2026-03-11T09:05:00Z",
        completed_at_utc="2026-03-11T09:06:00Z",
    )

    uploader_cfg = UploaderConfig(
        spool_dir=tmp_path,
        api_base_url="http://it.local",
        api_key="secret",
        retry=RetryConfig(max_attempts=2, initial_delay_s=1.0, max_delay_s=5.0, backoff_factor=2.0),
    )
    retention_cfg = SpoolRetentionConfig(
        enabled=True,
        max_age_days=45,
        protect_incomplete_runs=True,
        state_filename=".portal_upload_state.json",
    )
    client = TestClient(create_app(spool_dir=tmp_path, uploader_cfg=uploader_cfg, retention_cfg=retention_cfg))

    metrics = client.get("/metrics")
    assert metrics.status_code == 200
    metrics_payload = metrics.json()
    assert metrics_payload["runs_total"] == 2
    assert metrics_payload["events_total"] == 2
    assert metrics_payload["events_emitted_total"] == 2
    assert metrics_payload["count_a_to_b_total"] == 2
    assert metrics_payload["count_b_to_a_total"] == 0
    assert metrics_payload["delivery_state_counts"]["completed"] == 1
    assert metrics_payload["delivery_state_counts"]["pending"] == 1
    assert metrics_payload["sync_overview"]["status_cards"][3]["label"] == "Selesai"
    assert metrics_payload["sync_overview"]["status_cards"][3]["value"] == 1
    assert metrics_payload["lifecycle_status_counts"]["stopped"] == 2
    assert metrics_payload["latest_run"]["run_uid"] == "run_new"

    config = client.get("/config")
    assert config.status_code == 200
    config_payload = config.json()
    assert config_payload["uploader"]["enabled"] is True
    assert config_payload["uploader"]["api_key_configured"] is True
    assert "api_key" not in config_payload["uploader"]
    assert config_payload["service"] == {
        "exposure_mode": "loopback",
        "docs_enabled": True,
        "trusted_hosts": [],
    }
    assert config_payload["mutation_auth"] == {
        "enabled": False,
        "header_name": "X-API-Key",
    }
    assert config_payload["retention"] == {
        "enabled": True,
        "max_age_days": 45,
        "max_total_bytes": None,
        "min_free_bytes": None,
        "protect_incomplete_runs": True,
        "state_filename": ".portal_upload_state.json",
        "auto_run_interval_s": 0.0,
    }


def test_dashboard_payload_builds_real_hourly_trend(tmp_path) -> None:
    _write_run(
        tmp_path,
        day="2026-03-11",
        run_uid="run_a",
        started_at_utc="2026-03-11T00:00:00Z",
        occurred_at_utc="2026-03-11T00:15:00Z",
        occurred_at_local="2026-03-11T07:15:00+07:00",
        direction="A_TO_B",
    )
    _write_run(
        tmp_path,
        day="2026-03-11",
        run_uid="run_b",
        started_at_utc="2026-03-11T01:00:00Z",
        occurred_at_utc="2026-03-11T01:10:00Z",
        occurred_at_local="2026-03-11T08:10:00+07:00",
        direction="B_TO_A",
    )
    _write_run(
        tmp_path,
        day="2026-03-11",
        run_uid="run_c",
        started_at_utc="2026-03-11T01:00:00Z",
        occurred_at_utc="2026-03-11T01:40:00Z",
        occurred_at_local="2026-03-11T08:40:00+07:00",
        direction="A_TO_B",
    )

    app = create_app(spool_dir=tmp_path, review_db_path=tmp_path / "reviews.sqlite3")
    client = TestClient(app)
    review_resp = client.post(
        "/events/run_b_e1/review",
        json={"decision": "qualified_yes", "notes": "reviewed", "page_size": 25},
    )
    assert review_resp.status_code == 200

    trend = app.state.runtime.dashboard_payload(date_from="2026-03-11", date_to="2026-03-11")["trend"]
    assert trend["empty"] is False
    assert trend["bucket_mode"] == "hour"
    assert trend["bucket_hours"] == 24
    assert trend["time_basis_label"] == "Waktu (WIB)"
    assert trend["window_totals"] == {"a_to_b": 2, "b_to_a": 1, "pending": 2}
    assert trend["window_label"] == "Tanggal 2026-03-11 (WIB)"
    assert trend["buckets"][0]["label"] == "00:00"
    assert trend["buckets"][7]["label"] == "07:00"
    assert trend["buckets"][8]["label"] == "08:00"
    assert trend["buckets"][7]["a_to_b"] == 1
    assert trend["buckets"][7]["b_to_a"] == 0
    assert trend["buckets"][7]["pending"] == 1
    assert trend["buckets"][8]["a_to_b"] == 1
    assert trend["buckets"][8]["b_to_a"] == 1
    assert trend["buckets"][8]["pending"] == 1
    assert trend["series"][0]["path"].startswith("M")


def test_dashboard_payload_adapts_trend_buckets_for_longer_date_ranges(tmp_path) -> None:
    _write_run(
        tmp_path,
        day="2026-03-01",
        run_uid="run_mar_01",
        started_at_utc="2026-03-01T10:00:00Z",
        occurred_at_utc="2026-03-01T10:05:00Z",
    )
    _write_run(
        tmp_path,
        day="2026-03-05",
        run_uid="run_mar_05",
        started_at_utc="2026-03-05T10:00:00Z",
        occurred_at_utc="2026-03-05T10:05:00Z",
    )
    _write_run(
        tmp_path,
        day="2026-04-02",
        run_uid="run_apr_02",
        started_at_utc="2026-04-02T10:00:00Z",
        occurred_at_utc="2026-04-02T10:05:00Z",
    )

    app = create_app(spool_dir=tmp_path, review_db_path=tmp_path / "reviews.sqlite3")

    weekly_trend = app.state.runtime.dashboard_payload(date_from="2026-03-01", date_to="2026-03-07")["trend"]
    assert weekly_trend["bucket_mode"] == "day"
    assert weekly_trend["buckets"][0]["label"] == "03-01"
    assert weekly_trend["buckets"][4]["label"] == "03-05"

    monthly_trend = app.state.runtime.dashboard_payload(date_from="2026-03-01", date_to="2026-04-30")["trend"]
    assert monthly_trend["bucket_mode"] == "month"
    assert monthly_trend["buckets"][0]["label"] == "2026-03"
    assert monthly_trend["buckets"][1]["label"] == "2026-04"


def test_recent_runs_expose_sync_visibility_fields(tmp_path) -> None:
    _write_run(
        tmp_path,
        day="2026-03-11",
        run_uid="run_done",
        started_at_utc="2026-03-11T10:00:00Z",
        occurred_at_utc="2026-03-11T10:05:00Z",
        completed_at_utc="2026-03-11T10:06:00Z",
    )
    _write_run(
        tmp_path,
        day="2026-03-11",
        run_uid="run_inflight",
        started_at_utc="2026-03-11T10:10:00Z",
        occurred_at_utc="2026-03-11T10:15:00Z",
        in_progress_last_sync_at_utc="2026-03-11T10:16:00Z",
    )
    _write_run(
        tmp_path,
        day="2026-03-11",
        run_uid="run_failed",
        started_at_utc="2026-03-11T10:20:00Z",
        occurred_at_utc="2026-03-11T10:25:00Z",
        last_error="HTTP 503 on http://it.local/api/events/upsert: backend down",
        last_error_at_utc="2026-03-11T10:27:00Z",
    )
    _write_run(
        tmp_path,
        day="2026-03-11",
        run_uid="run_pending",
        started_at_utc="2026-03-11T10:30:00Z",
        occurred_at_utc="2026-03-11T10:35:00Z",
    )

    client = TestClient(create_app(spool_dir=tmp_path))

    runs = client.get("/runs/recent", params={"limit": 4})
    assert runs.status_code == 200
    by_uid = {item["run_uid"]: item for item in runs.json()["items"]}

    assert by_uid["run_done"]["delivery_state"] == "completed"
    assert by_uid["run_done"]["delivery_state_label"] == "Selesai"
    assert by_uid["run_done"]["delivery_state_pill_class"] == "yes"
    assert by_uid["run_done"]["last_sync_at_utc"] == "2026-03-11T10:06:00Z"
    assert by_uid["run_done"]["retry_recommended"] is False

    assert by_uid["run_inflight"]["delivery_state"] == "in_progress"
    assert by_uid["run_inflight"]["delivery_state_label"] == "Sedang dikirim"
    assert by_uid["run_inflight"]["last_sync_at_utc"] == "2026-03-11T10:16:00Z"
    assert by_uid["run_inflight"]["retry_recommended"] is False

    assert by_uid["run_failed"]["delivery_state"] == "failed"
    assert by_uid["run_failed"]["delivery_state_label"] == "Gagal"
    assert by_uid["run_failed"]["last_error_short"] == "Backend sync sedang bermasalah."
    assert by_uid["run_failed"]["retry_recommended"] is True

    assert by_uid["run_pending"]["delivery_state"] == "pending"
    assert by_uid["run_pending"]["delivery_state_label"] == "Pending"
    assert by_uid["run_pending"]["last_sync_at_utc"] is None
    assert by_uid["run_pending"]["retry_recommended"] is True


def test_retention_preview_and_run_endpoint_apply_policy(tmp_path) -> None:
    old_run_dir = _write_run(
        tmp_path,
        day="2025-11-01",
        run_uid="run_old",
        started_at_utc="2025-11-01T09:00:00Z",
        occurred_at_utc="2025-11-01T09:05:00Z",
        completed_at_utc="2025-11-01T09:06:00Z",
    )
    retention_cfg = SpoolRetentionConfig(
        enabled=True,
        max_age_days=30,
        protect_incomplete_runs=True,
        state_filename=".portal_upload_state.json",
    )
    client = TestClient(create_app(spool_dir=tmp_path, retention_cfg=retention_cfg))

    preview = client.get("/retention/preview")
    assert preview.status_code == 200
    preview_payload = preview.json()
    assert preview_payload["dry_run"] is True
    assert preview_payload["eligible_runs"] == 1
    assert preview_payload["deleted_runs"] == 0
    assert preview_payload["items"][0]["run_uid"] == "run_old"
    assert old_run_dir.exists()

    run_resp = client.post("/retention/run", json={"dry_run": False})
    assert run_resp.status_code == 200
    run_payload = run_resp.json()
    assert run_payload["dry_run"] is False
    assert run_payload["eligible_runs"] == 1
    assert run_payload["deleted_runs"] == 1
    assert not old_run_dir.exists()


def test_retention_run_endpoint_supports_pressure_overrides(tmp_path) -> None:
    _write_run(
        tmp_path,
        day="2026-03-01",
        run_uid="run_older",
        started_at_utc="2026-03-01T08:00:00Z",
        occurred_at_utc="2026-03-01T08:05:00Z",
        completed_at_utc="2026-03-01T08:06:00Z",
    )
    _write_run(
        tmp_path,
        day="2026-03-02",
        run_uid="run_newer",
        started_at_utc="2026-03-02T08:00:00Z",
        occurred_at_utc="2026-03-02T08:05:00Z",
        completed_at_utc="2026-03-02T08:06:00Z",
    )
    retention_cfg = SpoolRetentionConfig(
        enabled=True,
        max_age_days=365,
        protect_incomplete_runs=True,
        state_filename=".portal_upload_state.json",
    )
    app = create_app(spool_dir=tmp_path, retention_cfg=retention_cfg)
    baseline = app.state.runtime.run_retention(dry_run=True)
    client = TestClient(app)

    run_resp = client.post(
        "/retention/run",
        json={
            "dry_run": True,
            "max_total_bytes": int(baseline["total_runs_bytes"]) - 1,
        },
    )
    assert run_resp.status_code == 200
    payload = run_resp.json()
    assert payload["max_total_bytes"] == int(baseline["total_runs_bytes"]) - 1
    assert payload["eligible_runs"] == 1
    by_uid = {item["run_uid"]: item for item in payload["items"]}
    assert by_uid["run_older"]["selected_for_deletion"] is True
    assert by_uid["run_older"]["deletion_basis"] == "pressure"
    assert by_uid["run_newer"]["selected_for_deletion"] is False


def test_mutation_endpoints_require_api_key_when_configured(monkeypatch, tmp_path) -> None:
    _write_run(
        tmp_path,
        day="2026-03-11",
        run_uid="run_sync",
        started_at_utc="2026-03-11T10:00:00Z",
        occurred_at_utc="2026-03-11T10:05:00Z",
    )

    uploader_cfg = UploaderConfig(
        spool_dir=tmp_path,
        api_base_url="http://it.local",
        api_key="secret",
        retry=RetryConfig(max_attempts=1, initial_delay_s=0.0, max_delay_s=0.0, backoff_factor=1.0),
    )
    client = TestClient(
        create_app(
            spool_dir=tmp_path,
            uploader_cfg=uploader_cfg,
            mutation_auth_cfg=MutationAuthConfig(api_key="edge-local-secret"),
        )
    )

    def _fake_process_pending_runs(cfg, *, force: bool = False, dry_run: bool = False, max_runs=None):
        _ = cfg
        _ = force
        _ = dry_run
        _ = max_runs
        return SyncSummary(discovered_runs=1, completed_runs=1, skipped_runs=0, failed_runs=0)

    monkeypatch.setattr(api_module, "process_pending_runs", _fake_process_pending_runs)

    status_resp = client.get("/status")
    assert status_resp.status_code == 200
    assert status_resp.json()["mutation_auth_enabled"] is True

    unauthorized = client.post("/sync/retry", json={"dry_run": True})
    assert unauthorized.status_code == 401
    assert unauthorized.json()["detail"] == "invalid_api_key"

    wrong_key = client.post("/sync/retry", json={"dry_run": True}, headers={"X-API-Key": "wrong"})
    assert wrong_key.status_code == 401

    authorized = client.post("/sync/retry", json={"dry_run": True}, headers={"X-API-Key": "edge-local-secret"})
    assert authorized.status_code == 200
    assert authorized.json()["summary"]["completed_runs"] == 1


def test_service_guardrails_require_mutation_key_for_non_loopback_host() -> None:
    service_module._validate_mutation_auth_guardrails(
        "127.0.0.1",
        MutationAuthConfig(api_key="", header_name="X-API-Key"),
    )
    service_module._validate_mutation_auth_guardrails(
        "0.0.0.0",
        MutationAuthConfig(api_key="edge-local-secret", header_name="X-API-Key"),
    )

    try:
        service_module._validate_mutation_auth_guardrails(
            "0.0.0.0",
            MutationAuthConfig(api_key="", header_name="X-API-Key"),
        )
    except SystemExit as exc:
        assert "non-loopback host" in str(exc)
    else:
        raise AssertionError("Expected SystemExit for non-loopback host without mutation API key")


def test_parse_service_trusted_hosts_supports_repeated_and_comma_separated_values() -> None:
    parsed = service_module._parse_service_trusted_hosts(
        [
            "localhost, 127.0.0.1",
            "edge.local",
            "127.0.0.1, 10.0.0.25",
        ]
    )

    assert parsed == [
        "localhost",
        "127.0.0.1",
        "edge.local",
        "10.0.0.25",
    ]


def test_service_guardrails_require_lan_mode_ui_auth_docs_off_and_trusted_hosts() -> None:
    base_service_cfg = ServiceConfig(
        exposure_mode="lan",
        enable_docs=False,
        trusted_hosts=["10.0.0.25", "localhost"],
    )
    base_mutation_cfg = MutationAuthConfig(api_key="edge-local-secret", header_name="X-API-Key")
    base_ui_cfg = UiAuthConfig(username="admin", password="secret")

    service_module._validate_service_guardrails(
        "0.0.0.0",
        base_service_cfg,
        base_mutation_cfg,
        base_ui_cfg,
    )

    for broken_cfg, broken_mutation, broken_ui, expected_text in (
        (
            ServiceConfig(exposure_mode="loopback", enable_docs=False, trusted_hosts=["10.0.0.25"]),
            base_mutation_cfg,
            base_ui_cfg,
            "service.exposure_mode='loopback'",
        ),
        (
            ServiceConfig(exposure_mode="lan", enable_docs=True, trusted_hosts=["10.0.0.25"]),
            base_mutation_cfg,
            base_ui_cfg,
            "docs are enabled",
        ),
        (
            ServiceConfig(exposure_mode="lan", enable_docs=False, trusted_hosts=["10.0.0.25"]),
            MutationAuthConfig(api_key="", header_name="X-API-Key"),
            base_ui_cfg,
            "mutation endpoint protection",
        ),
        (
            ServiceConfig(exposure_mode="lan", enable_docs=False, trusted_hosts=["10.0.0.25"]),
            base_mutation_cfg,
            UiAuthConfig(username="admin", password=""),
            "without UI authentication",
        ),
        (
            ServiceConfig(exposure_mode="lan", enable_docs=False, trusted_hosts=[]),
            base_mutation_cfg,
            base_ui_cfg,
            "without configured trusted hosts",
        ),
        (
            ServiceConfig(exposure_mode="lan", enable_docs=False, trusted_hosts=["*"]),
            base_mutation_cfg,
            base_ui_cfg,
            "wildcard trusted hosts",
        ),
    ):
        try:
            service_module._validate_service_guardrails(
                "0.0.0.0",
                broken_cfg,
                broken_mutation,
                broken_ui,
            )
        except SystemExit as exc:
            assert expected_text in str(exc)
        else:
            raise AssertionError(f"Expected SystemExit containing {expected_text!r}")


def test_create_app_can_disable_docs_endpoints(tmp_path) -> None:
    client = TestClient(
        create_app(
            spool_dir=tmp_path,
            service_cfg=ServiceConfig(enable_docs=False),
        )
    )

    docs = client.get("/docs")
    assert docs.status_code == 404

    redoc = client.get("/redoc")
    assert redoc.status_code == 404

    openapi = client.get("/openapi.json")
    assert openapi.status_code == 404


def test_create_app_enforces_trusted_hosts(tmp_path) -> None:
    client = TestClient(
        create_app(
            spool_dir=tmp_path,
            service_cfg=ServiceConfig(trusted_hosts=["edge.local"]),
        ),
        base_url="http://edge.local",
    )

    good = client.get("/healthz")
    assert good.status_code == 200

    bad = client.get("/healthz", headers={"host": "unexpected.local"})
    assert bad.status_code == 400


def test_sync_endpoints_use_existing_uploader_flow(monkeypatch, tmp_path) -> None:
    _write_run(
        tmp_path,
        day="2026-03-11",
        run_uid="run_sync",
        started_at_utc="2026-03-11T10:00:00Z",
        occurred_at_utc="2026-03-11T10:05:00Z",
    )

    uploader_cfg = UploaderConfig(
        spool_dir=tmp_path,
        api_base_url="http://it.local",
        api_key="secret",
        retry=RetryConfig(max_attempts=1, initial_delay_s=0.0, max_delay_s=0.0, backoff_factor=1.0),
    )
    client = TestClient(create_app(spool_dir=tmp_path, uploader_cfg=uploader_cfg))

    recorded: dict[str, object] = {}

    def _fake_process_pending_runs(cfg, *, force: bool = False, dry_run: bool = False, max_runs=None):
        recorded["pending"] = {
            "spool_dir": str(cfg.spool_dir),
            "force": bool(force),
            "dry_run": bool(dry_run),
            "max_runs": max_runs,
        }
        return SyncSummary(discovered_runs=1, completed_runs=1, skipped_runs=0, failed_runs=0)

    class _FakeClient:
        def __init__(self, *, base_url: str, api_key: str, timeout_s: float) -> None:
            recorded["client"] = {
                "base_url": base_url,
                "api_key": api_key,
                "timeout_s": timeout_s,
            }

    def _fake_process_single_run(run_dir, *, cfg, client, force: bool = False, dry_run: bool = False):
        recorded["single"] = {
            "run_dir": str(run_dir),
            "force": bool(force),
            "dry_run": bool(dry_run),
            "api_base_url": cfg.api_base_url,
            "client_type": type(client).__name__,
        }
        return "completed"

    monkeypatch.setattr(api_module, "process_pending_runs", _fake_process_pending_runs)
    monkeypatch.setattr(api_module, "DeliveryApiClient", _FakeClient)
    monkeypatch.setattr(api_module, "process_single_run", _fake_process_single_run)

    retry_resp = client.post("/sync/retry", json={"force": True, "dry_run": True, "max_runs": 3})
    assert retry_resp.status_code == 200
    assert retry_resp.json()["summary"]["completed_runs"] == 1
    assert recorded["pending"] == {
        "spool_dir": str(tmp_path),
        "force": True,
        "dry_run": True,
        "max_runs": 3,
    }

    single_resp = client.post("/sync/run/run_sync", json={"force": False, "dry_run": True})
    assert single_resp.status_code == 200
    assert single_resp.json()["status"] == "completed"
    assert recorded["single"] == {
        "run_dir": str(tmp_path / "2026-03-11" / "run_sync"),
        "force": False,
        "dry_run": True,
        "api_base_url": "http://it.local",
        "client_type": "_FakeClient",
    }


def test_review_api_and_event_detail_include_sqlite_backed_review_state(tmp_path) -> None:
    _write_run(
        tmp_path,
        day="2026-03-11",
        run_uid="run_ui",
        started_at_utc="2026-03-11T10:00:00Z",
        occurred_at_utc="2026-03-11T10:05:00Z",
    )

    client = TestClient(create_app(spool_dir=tmp_path, review_db_path=tmp_path / "reviews.sqlite3"))

    queue_before = client.get("/review/queue")
    assert queue_before.status_code == 200
    assert queue_before.json()["queue_total"] == 1
    assert queue_before.json()["current"]["review_status"] == "pending"

    review_resp = client.post(
        "/events/run_ui_e1/review",
        json={
            "decision": "qualified_yes",
            "reviewed_class": "pickup",
            "notes": "matches target vehicle",
        },
    )
    assert review_resp.status_code == 200
    review_payload = review_resp.json()
    assert review_payload["review"]["decision"] == "qualified_yes"
    assert review_payload["review"]["reviewed_class"] == "pickup"
    assert review_payload["review"]["notes"] == "matches target vehicle"

    event_resp = client.get("/events/run_ui_e1")
    assert event_resp.status_code == 200
    event_payload = event_resp.json()
    assert event_payload["review_status"] == "qualified_yes"
    assert event_payload["model_class_name"] == "truck"
    assert event_payload["reviewed_class_name"] == "pickup"
    assert event_payload["effective_class_name"] == "pickup"
    assert event_payload["review"]["notes"] == "matches target vehicle"
    assert event_payload["timeline"][-1]["description"].startswith("Direview sebagai Diterima")


def test_review_api_keeps_model_class_as_effective_when_override_matches_prediction(tmp_path) -> None:
    _write_run(
        tmp_path,
        day="2026-03-11",
        run_uid="run_same_class",
        started_at_utc="2026-03-11T10:00:00Z",
        occurred_at_utc="2026-03-11T10:05:00Z",
    )

    client = TestClient(create_app(spool_dir=tmp_path, review_db_path=tmp_path / "reviews.sqlite3"))

    review_resp = client.post(
        "/events/run_same_class_e1/review",
        json={
            "decision": "qualified_yes",
            "reviewed_class": "truck",
            "notes": "same as detected class",
        },
    )
    assert review_resp.status_code == 200
    payload = review_resp.json()
    assert payload["review"]["reviewed_class"] is None

    event_payload = client.get("/events/run_same_class_e1").json()
    assert event_payload["model_class_name"] == "truck"
    assert event_payload["reviewed_class_name"] is None
    assert event_payload["effective_class_name"] == "truck"


def test_review_api_does_not_use_corrected_class_as_effective_for_rejected_events(tmp_path) -> None:
    _write_run(
        tmp_path,
        day="2026-03-11",
        run_uid="run_rejected",
        started_at_utc="2026-03-11T10:00:00Z",
        occurred_at_utc="2026-03-11T10:05:00Z",
    )

    client = TestClient(create_app(spool_dir=tmp_path, review_db_path=tmp_path / "reviews.sqlite3"))

    review_resp = client.post(
        "/events/run_rejected_e1/review",
        json={
            "decision": "qualified_no",
            "reviewed_class": "pickup",
            "reject_reason": "not_relevant",
            "notes": "wrong target for operations",
        },
    )
    assert review_resp.status_code == 200
    payload = review_resp.json()
    assert payload["review"]["reviewed_class"] == "pickup"

    event_payload = client.get("/events/run_rejected_e1").json()
    assert event_payload["review_status"] == "qualified_no"
    assert event_payload["model_class_name"] == "truck"
    assert event_payload["reviewed_class_name"] == "pickup"
    assert event_payload["effective_class_name"] is None


def test_review_queue_page_uses_effective_class_after_yes_review(tmp_path) -> None:
    _write_run(
        tmp_path,
        day="2026-03-11",
        run_uid="run_queue_class",
        started_at_utc="2026-03-11T10:00:00Z",
        occurred_at_utc="2026-03-11T10:05:00Z",
    )

    client = TestClient(create_app(spool_dir=tmp_path, review_db_path=tmp_path / "reviews.sqlite3"))
    review_resp = client.post(
        "/events/run_queue_class_e1/review",
        json={
            "decision": "qualified_yes",
            "reviewed_class": "pickup",
            "notes": "corrected for operations",
        },
    )
    assert review_resp.status_code == 200

    review_page = client.get("/ui/review?status=all&page=1&page_size=25")
    assert review_page.status_code == 200
    assert "pickup" in review_page.text
    assert "model truck" in review_page.text
    assert "review pickup" in review_page.text


def test_review_store_migrates_existing_db_to_support_reviewed_class(tmp_path) -> None:
    db_path = tmp_path / "reviews.sqlite3"
    conn = sqlite3.connect(db_path)
    conn.execute(
        """
        CREATE TABLE event_reviews (
            event_uid TEXT PRIMARY KEY,
            run_uid TEXT,
            site_id TEXT,
            camera_id TEXT,
            decision TEXT NOT NULL,
            notes TEXT NOT NULL DEFAULT '',
            created_at_utc TEXT NOT NULL,
            updated_at_utc TEXT NOT NULL
        )
        """
    )
    conn.commit()
    conn.close()

    store = ReviewStore(db_path)
    record = store.save_review(
        event_uid="event_1",
        run_uid="run_1",
        site_id="site_a",
        camera_id="cam_01",
        decision="qualified_yes",
        reviewed_class="pickup",
        notes="migrated db should accept corrected class",
        now_utc="2026-03-11T10:05:00Z",
    )

    assert record.reviewed_class == "pickup"
    loaded = store.get_review("event_1")
    assert loaded is not None
    assert loaded.reviewed_class == "pickup"


def test_event_detail_page_exposes_reviewed_class_input_and_suggestions(tmp_path) -> None:
    _write_run(
        tmp_path,
        day="2026-03-11",
        run_uid="run_detail_class",
        started_at_utc="2026-03-11T10:00:00Z",
        occurred_at_utc="2026-03-11T10:05:00Z",
    )
    _write_run(
        tmp_path,
        day="2026-03-11",
        run_uid="run_detail_pickup",
        started_at_utc="2026-03-11T10:10:00Z",
        occurred_at_utc="2026-03-11T10:15:00Z",
        class_name="pickup",
    )

    client = TestClient(create_app(spool_dir=tmp_path, review_db_path=tmp_path / "reviews.sqlite3"))
    detail = client.get("/ui/events/run_detail_class_e1?status=pending&page=1&page_size=25")

    assert detail.status_code == 200
    assert '<select' in detail.text
    assert 'name="reviewed_class"' in detail.text
    assert 'name="reject_reason"' in detail.text
    assert 'data-review-reject-picker' in detail.text
    assert 'data-review-reject-toggle' in detail.text
    assert 'data-review-reject-submit' in detail.text
    assert 'data-last-review-notice' in detail.text
    assert "Ubah keputusan terakhir" in detail.text
    assert 'aria-expanded="false"' in detail.text
    assert "Klik alasan atau tekan 1–6" in detail.text
    assert "Tolak data ini" not in detail.text
    reject_toggle_at = detail.text.index('data-review-reject-toggle')
    reject_toggle_start = detail.text.rfind("<button", 0, reject_toggle_at)
    reject_toggle_end = detail.text.find(">", reject_toggle_at)
    reject_toggle_tag = detail.text[reject_toggle_start:reject_toggle_end]
    assert 'type="button"' in reject_toggle_tag
    assert 'name="decision"' not in reject_toggle_tag
    assert "Terima" in detail.text
    assert "Tolak" in detail.text
    assert detail.text.index("Terima") < detail.text.index('name="reviewed_class"')
    assert detail.text.index("Tolak") < detail.text.index('name="reject_reason"')
    assert "Tetap gunakan tipe terdeteksi" in detail.text
    assert "False positive / bukan kendaraan target" in detail.text
    assert '<option value="pickup"' in detail.text


def test_event_detail_uses_full_run_taxonomy_before_classes_are_predicted(tmp_path) -> None:
    _write_run(
        tmp_path,
        day="2026-03-11",
        run_uid="run_full_taxonomy",
        started_at_utc="2026-03-11T10:00:00Z",
        occurred_at_utc="2026-03-11T10:05:00Z",
        class_name="pickup",
        run_class_names={
            0: "container 20ft",
            1: "container 40ft",
            2: "trailer",
            3: "tronton",
            4: "engkel",
            5: "double engkel",
            6: "pickup",
        },
    )

    client = TestClient(create_app(spool_dir=tmp_path, review_db_path=tmp_path / "reviews.sqlite3"))
    detail = client.get("/ui/events/run_full_taxonomy_e1?status=pending&page=1&page_size=25")

    assert detail.status_code == 200
    for class_name in (
        "container 20ft",
        "container 40ft",
        "trailer",
        "tronton",
        "engkel",
        "double engkel",
        "pickup",
    ):
        assert f'<option value="{class_name}"' in detail.text


def test_event_detail_reject_reason_is_saved_in_existing_notes(tmp_path) -> None:
    _write_run(
        tmp_path,
        day="2026-03-11",
        run_uid="run_reject_reason",
        started_at_utc="2026-03-11T10:00:00Z",
        occurred_at_utc="2026-03-11T10:05:00Z",
    )

    client = TestClient(create_app(spool_dir=tmp_path, review_db_path=tmp_path / "reviews.sqlite3"))

    response = client.post(
        "/ui/events/run_reject_reason_e1/review",
        data={
            "decision": "qualified_no",
            "reject_reason": "false_positive",
            "notes": "shadow on the lane marker",
            "status_filter": "pending",
            "page": "1",
            "page_size": "25",
        },
        follow_redirects=False,
    )

    assert response.status_code == 303
    payload = client.get("/events/run_reject_reason_e1").json()
    assert payload["review"]["decision"] == "qualified_no"
    assert payload["review"]["notes"] == (
        "Alasan tolak: False positive / bukan kendaraan target\n"
        "shadow on the lane marker"
    )

    detail = client.get("/ui/events/run_reject_reason_e1?status=all&page=1&page_size=25")
    assert detail.status_code == 200
    assert 'value="false_positive"' in detail.text
    assert 'data-review-reject-reason' in detail.text
    assert 'checked' in detail.text
    assert "shadow on the lane marker" in detail.text


def test_rejected_review_requires_known_reason_for_api_and_form(tmp_path) -> None:
    _write_run(
        tmp_path,
        day="2026-03-11",
        run_uid="run_required_reject_reason",
        started_at_utc="2026-03-11T10:00:00Z",
        occurred_at_utc="2026-03-11T10:05:00Z",
    )
    client = TestClient(create_app(spool_dir=tmp_path, review_db_path=tmp_path / "reviews.sqlite3"))
    review_url = "/events/run_required_reject_reason_e1/review"

    missing_api_reason = client.post(review_url, json={"decision": "qualified_no"})
    assert missing_api_reason.status_code == 422
    assert missing_api_reason.json()["detail"] == (
        "reject_reason is required when decision is qualified_no"
    )

    invalid_api_reason = client.post(
        review_url,
        json={"decision": "qualified_no", "reject_reason": "quick_default"},
    )
    assert invalid_api_reason.status_code == 422
    assert invalid_api_reason.json()["detail"] == "reject_reason is invalid"

    missing_form_reason = client.post(
        "/ui/events/run_required_reject_reason_e1/review",
        data={"decision": "qualified_no"},
    )
    assert missing_form_reason.status_code == 422
    assert client.get("/events/run_required_reject_reason_e1").json()["review"] is None


def test_review_api_returns_next_pending_event_for_detail_flow(tmp_path) -> None:
    _write_run(
        tmp_path,
        day="2026-03-11",
        run_uid="run_old",
        started_at_utc="2026-03-11T10:00:00Z",
        occurred_at_utc="2026-03-11T10:05:00Z",
    )
    _write_run(
        tmp_path,
        day="2026-03-11",
        run_uid="run_new",
        started_at_utc="2026-03-11T10:10:00Z",
        occurred_at_utc="2026-03-11T10:15:00Z",
    )

    client = TestClient(create_app(spool_dir=tmp_path, review_db_path=tmp_path / "reviews.sqlite3"))

    review_resp = client.post(
        "/events/run_new_e1/review",
        json={
            "decision": "qualified_yes",
            "notes": "move forward",
            "camera_id": "cam_01",
            "status_filter": "pending",
            "page": 1,
            "page_size": 25,
        },
    )
    assert review_resp.status_code == 200
    payload = review_resp.json()
    assert payload["next_event_uid"] == "run_old_e1"
    assert payload["next_detail_url"] == "/ui/events/run_old_e1?camera_id=cam_01&status=pending&page=1&page_size=25"
    assert payload["queue_page_url"] == "/ui/review?camera_id=cam_01&status=pending&page=1&page_size=25"
    assert payload["next_pending_detail_url"] == "/ui/events/run_old_e1?camera_id=cam_01&status=pending&page=1&page_size=25"
    assert payload["next_pending_queue_url"] == "/ui/review?camera_id=cam_01&status=pending&page=1&page_size=25"

    detail = client.get("/ui/events/run_new_e1?camera_id=cam_01&status=pending&page=1&page_size=25")
    assert detail.status_code == 200
    assert "data-status-filter=\"pending\"" in detail.text
    assert "data-page=\"1\"" in detail.text
    assert "data-pending-queue-url=\"/ui/review?camera_id=cam_01&amp;status=pending&amp;page=1&amp;page_size=25\"" in detail.text


def test_review_save_does_not_enrich_complete_backlog(monkeypatch, tmp_path) -> None:
    for index in range(40):
        _write_run(
            tmp_path,
            day="2026-03-11",
            run_uid=f"run_save_bounded_{index:02d}",
            started_at_utc=f"2026-03-11T10:{index:02d}:00Z",
            occurred_at_utc=f"2026-03-11T10:{index:02d}:30Z",
        )

    app = create_app(
        spool_dir=tmp_path,
        review_db_path=tmp_path / "reviews.sqlite3",
    )
    runtime = app.state.runtime
    original_attach = runtime._attach_review_data
    enriched_batch_sizes = []

    def record_attach(items, *, review_map=None):
        enriched_batch_sizes.append(len(items))
        return original_attach(items, review_map=review_map)

    monkeypatch.setattr(runtime, "_attach_review_data", record_attach)
    client = TestClient(app)

    response = client.post(
        "/events/run_save_bounded_39_e1/review",
        json={
            "decision": "qualified_yes",
            "camera_id": "cam_01",
            "status_filter": "pending",
            "page": 3,
            "page_size": 15,
        },
    )

    assert response.status_code == 200
    assert response.json()["next_event_uid"] == "run_save_bounded_00_e1"
    assert enriched_batch_sizes
    assert max(enriched_batch_sizes) <= 16


def test_review_api_uses_global_pending_queue_when_detail_is_not_camera_filtered(tmp_path) -> None:
    _write_run(
        tmp_path,
        day="2026-03-11",
        run_uid="run_cam_a",
        started_at_utc="2026-03-11T10:00:00Z",
        occurred_at_utc="2026-03-11T10:05:00Z",
        camera_id="cam_01",
    )
    _write_run(
        tmp_path,
        day="2026-03-11",
        run_uid="run_cam_b",
        started_at_utc="2026-03-11T10:10:00Z",
        occurred_at_utc="2026-03-11T10:15:00Z",
        camera_id="cam_02",
    )

    client = TestClient(create_app(spool_dir=tmp_path, review_db_path=tmp_path / "reviews.sqlite3"))

    review_resp = client.post(
        "/events/run_cam_b_e1/review",
        json={
            "decision": "qualified_yes",
            "notes": "continue globally",
            "camera_id": None,
            "status_filter": "pending",
            "page": 1,
            "page_size": 25,
        },
    )
    assert review_resp.status_code == 200
    payload = review_resp.json()
    assert payload["next_event_uid"] == "run_cam_a_e1"
    assert payload["next_detail_url"] == "/ui/events/run_cam_a_e1?status=pending&page=1&page_size=25"
    assert payload["queue_page_url"] == "/ui/review?status=pending&page=1&page_size=25"
    assert payload["next_pending_detail_url"] == "/ui/events/run_cam_a_e1?status=pending&page=1&page_size=25"
    assert payload["next_pending_queue_url"] == "/ui/review?status=pending&page=1&page_size=25"

    detail = client.get("/ui/events/run_cam_b_e1?status=pending&page=1&page_size=25")
    assert detail.status_code == 200
    assert "data-camera-id=\"\"" in detail.text
    assert "data-pending-queue-url=\"/ui/review?status=pending&amp;page=1&amp;page_size=25\"" in detail.text


def test_review_api_preserves_all_queue_continuation_path(tmp_path) -> None:
    _write_run(
        tmp_path,
        day="2026-03-11",
        run_uid="run_old",
        started_at_utc="2026-03-11T10:00:00Z",
        occurred_at_utc="2026-03-11T10:05:00Z",
    )
    _write_run(
        tmp_path,
        day="2026-03-11",
        run_uid="run_new",
        started_at_utc="2026-03-11T10:10:00Z",
        occurred_at_utc="2026-03-11T10:15:00Z",
    )

    client = TestClient(create_app(spool_dir=tmp_path, review_db_path=tmp_path / "reviews.sqlite3"))

    review_resp = client.post(
        "/events/run_new_e1/review",
        json={
            "decision": "qualified_yes",
            "notes": "continue in all queue",
            "camera_id": "cam_01",
            "status_filter": "all",
            "page": 1,
            "page_size": 25,
        },
    )
    assert review_resp.status_code == 200
    payload = review_resp.json()
    assert payload["next_detail_url"] == "/ui/events/run_old_e1?camera_id=cam_01&status=all&page=1&page_size=25"
    assert payload["queue_page_url"] == "/ui/review?camera_id=cam_01&status=all&page=1&page_size=25"
    assert payload["next_pending_detail_url"] == "/ui/events/run_old_e1?camera_id=cam_01&status=pending&page=1&page_size=25"


def test_event_detail_form_submit_redirects_to_next_detail(tmp_path) -> None:
    _write_run(
        tmp_path,
        day="2026-03-11",
        run_uid="run_old",
        started_at_utc="2026-03-11T10:00:00Z",
        occurred_at_utc="2026-03-11T10:05:00Z",
    )
    _write_run(
        tmp_path,
        day="2026-03-11",
        run_uid="run_new",
        started_at_utc="2026-03-11T10:10:00Z",
        occurred_at_utc="2026-03-11T10:15:00Z",
    )

    client = TestClient(create_app(spool_dir=tmp_path, review_db_path=tmp_path / "reviews.sqlite3"))

    response = client.post(
        "/ui/events/run_new_e1/review",
        data={
            "decision": "qualified_yes",
            "notes": "use redirect flow",
            "camera_id": "cam_01",
            "status_filter": "pending",
            "page": "1",
            "page_size": "25",
        },
        follow_redirects=False,
    )
    assert response.status_code == 303
    assert response.headers["location"] == "/ui/events/run_old_e1?camera_id=cam_01&status=pending&page=1&page_size=25"


def test_review_queue_date_filter_preserves_state_in_detail_links(tmp_path) -> None:
    _write_run(
        tmp_path,
        day="2026-03-10",
        run_uid="run_old_day",
        started_at_utc="2026-03-10T10:00:00Z",
        occurred_at_utc="2026-03-10T10:05:00Z",
    )
    _write_run(
        tmp_path,
        day="2026-03-11",
        run_uid="run_target_day",
        started_at_utc="2026-03-11T10:00:00Z",
        occurred_at_utc="2026-03-11T10:05:00Z",
    )

    client = TestClient(create_app(spool_dir=tmp_path, review_db_path=tmp_path / "reviews.sqlite3"))

    queue = client.get(
        "/review/queue",
        params={
            "status": "pending",
            "camera_id": "cam_01",
            "page": 1,
            "page_size": 25,
            "date_from": "2026-03-11",
            "date_to": "2026-03-11",
        },
    )
    assert queue.status_code == 200
    payload = queue.json()
    assert payload["queue_total"] == 1
    assert payload["items"][0]["event_uid"] == "run_target_day_e1"
    assert payload["items"][0]["detail_url"] == "/ui/events/run_target_day_e1?camera_id=cam_01&status=pending&page=1&page_size=25&date_from=2026-03-11&date_to=2026-03-11"

    detail = client.get("/ui/events/run_target_day_e1?camera_id=cam_01&status=pending&page=1&page_size=25&date_from=2026-03-11&date_to=2026-03-11")
    assert detail.status_code == 200
    assert "value=\"2026-03-11\"" in detail.text
    assert "/ui/review?camera_id=cam_01&amp;status=pending&amp;event_uid=run_target_day_e1&amp;page=1&amp;page_size=25&amp;date_from=2026-03-11&amp;date_to=2026-03-11" in detail.text


def test_review_submit_keeps_date_range_and_stays_within_filtered_queue(tmp_path) -> None:
    _write_run(
        tmp_path,
        day="2026-03-10",
        run_uid="run_outside_range",
        started_at_utc="2026-03-10T10:00:00Z",
        occurred_at_utc="2026-03-10T10:05:00Z",
    )
    _write_run(
        tmp_path,
        day="2026-03-11",
        run_uid="run_in_range_old",
        started_at_utc="2026-03-11T10:00:00Z",
        occurred_at_utc="2026-03-11T10:05:00Z",
    )
    _write_run(
        tmp_path,
        day="2026-03-11",
        run_uid="run_in_range_new",
        started_at_utc="2026-03-11T10:10:00Z",
        occurred_at_utc="2026-03-11T10:15:00Z",
    )

    client = TestClient(create_app(spool_dir=tmp_path, review_db_path=tmp_path / "reviews.sqlite3"))

    review_resp = client.post(
        "/events/run_in_range_new_e1/review",
        json={
            "decision": "qualified_yes",
            "notes": "filtered review",
            "camera_id": "cam_01",
            "status_filter": "pending",
            "page": 1,
            "page_size": 25,
            "date_from": "2026-03-11",
            "date_to": "2026-03-11",
        },
    )
    assert review_resp.status_code == 200
    payload = review_resp.json()
    assert payload["next_event_uid"] == "run_in_range_old_e1"
    assert payload["next_detail_url"] == "/ui/events/run_in_range_old_e1?camera_id=cam_01&status=pending&page=1&page_size=25&date_from=2026-03-11&date_to=2026-03-11"
    assert payload["next_pending_detail_url"] == "/ui/events/run_in_range_old_e1?camera_id=cam_01&status=pending&page=1&page_size=25&date_from=2026-03-11&date_to=2026-03-11"

    response = client.post(
        "/ui/events/run_in_range_new_e1/review",
        data={
            "decision": "qualified_yes",
            "notes": "use redirect flow",
            "camera_id": "cam_01",
            "status_filter": "pending",
            "page": "1",
            "page_size": "25",
            "date_from": "2026-03-11",
            "date_to": "2026-03-11",
        },
        follow_redirects=False,
    )
    assert response.status_code == 303
    assert response.headers["location"] == "/ui/events/run_in_range_old_e1?camera_id=cam_01&status=pending&page=1&page_size=25&date_from=2026-03-11&date_to=2026-03-11"


def test_ui_pages_render_dashboard_queue_and_detail(tmp_path) -> None:
    _write_run(
        tmp_path,
        day="2026-03-11",
        run_uid="run_ui",
        started_at_utc="2026-03-11T10:00:00Z",
        occurred_at_utc="2026-03-11T10:05:00Z",
        occurred_at_local="2026-03-11T17:05:00+07:00",
    )
    _write_run(
        tmp_path,
        day="2026-03-11",
        run_uid="run_sync_failed",
        started_at_utc="2026-03-11T10:10:00Z",
        occurred_at_utc="2026-03-11T10:15:00Z",
        last_error="network error on http://it.local/api/events/upsert: connection refused",
        last_error_at_utc="2026-03-11T10:16:00Z",
    )
    _write_run(
        tmp_path,
        day="2026-03-11",
        run_uid="run_sync_live",
        started_at_utc="2026-03-11T10:20:00Z",
        occurred_at_utc="2026-03-11T10:25:00Z",
        in_progress_last_sync_at_utc="2026-03-11T10:26:00Z",
    )
    client = TestClient(create_app(spool_dir=tmp_path, review_db_path=tmp_path / "reviews.sqlite3"))

    dashboard = client.get("/ui/dashboard?date_from=2026-03-11&date_to=2026-03-11")
    assert dashboard.status_code == 200
    assert "Traffic Monitoring Dashboard" in dashboard.text
    assert "Event terbaru" in dashboard.text
    assert "Pengiriman Data" in dashboard.text
    assert "Sedang dikirim" in dashboard.text
    assert "Gagal" in dashboard.text
    assert "Koneksi backend gagal." in dashboard.text
    assert "Sync terakhir" in dashboard.text
    assert "Traffic trend chart from local spool data" in dashboard.text
    assert "placeholder chart" not in dashboard.text
    assert "Waste Paper" in dashboard.text
    assert "Finished Goods" in dashboard.text
    assert "A_TO_B" not in dashboard.text
    assert "B_TO_A" not in dashboard.text
    assert "value=\"2026-03-11\"" in dashboard.text
    assert "/ui/review?status=pending&amp;date_from=2026-03-11&amp;date_to=2026-03-11" in dashboard.text

    review = client.get("/ui/review?camera_id=cam_01&status=pending&page=1&page_size=25&date_from=2026-03-11&date_to=2026-03-11")
    assert review.status_code == 200
    assert "Antrian Review" in review.text
    assert "Periksa data kendaraan" in review.text
    assert "SQLite" not in review.text
    assert "run_ui_e1" in review.text
    assert "17:05 WIB" in review.text
    assert "Waste Paper" in review.text
    assert "A_TO_B" not in review.text
    assert ">Aksi<" in review.text
    assert ">Buka<" in review.text
    assert "queue-action-cell" in review.text
    assert "/ui/events/run_ui_e1?camera_id=cam_01&amp;status=pending&amp;page=1&amp;page_size=25&amp;date_from=2026-03-11&amp;date_to=2026-03-11" in review.text

    detail = client.get("/ui/events/run_ui_e1?camera_id=cam_01&status=pending&page=1&page_size=25&date_from=2026-03-11&date_to=2026-03-11")
    assert detail.status_code == 200
    assert "Event Detail" in detail.text
    assert "run_ui_e1" in detail.text
    assert "2026-03-11 17:05 WIB" in detail.text
    assert "Pending" in detail.text
    assert "Waste Paper" in detail.text
    assert "A_TO_B" not in detail.text
    assert "/ui/review?camera_id=cam_01&amp;status=pending&amp;event_uid=run_ui_e1&amp;page=1&amp;page_size=25&amp;date_from=2026-03-11&amp;date_to=2026-03-11" in detail.text

    css = client.get("/ui-static/app.css")
    assert css.status_code == 200
    assert ".login-shell" in css.text

    js = client.get("/ui-static/app.js")
    assert js.status_code == 200
    assert "initReviewActions" in js.text

    recent_events = client.get("/events/recent").json()
    assert recent_events["items"][0]["direction"] == "A_TO_B"


def test_review_queue_paginates_and_preserves_page_state(tmp_path) -> None:
    for index in range(30):
        _write_run(
            tmp_path,
            day="2026-03-11",
            run_uid=f"run_page_{index:02d}",
            started_at_utc=f"2026-03-11T10:{index:02d}:00Z",
            occurred_at_utc=f"2026-03-11T10:{index:02d}:30Z",
        )

    client = TestClient(create_app(spool_dir=tmp_path, review_db_path=tmp_path / "reviews.sqlite3"))

    queue = client.get("/review/queue", params={"status": "pending", "page": 2, "page_size": 15})
    assert queue.status_code == 200
    payload = queue.json()
    assert payload["queue_total"] == 30
    assert payload["page_item_count"] == 15
    assert payload["pagination"]["current_page"] == 2
    assert payload["pagination"]["total_pages"] == 2
    assert payload["pagination"]["start_item"] == 16
    assert payload["pagination"]["end_item"] == 30
    assert len(payload["items"]) == 15
    assert payload["items"][0]["detail_url"].endswith("&page=2&page_size=15")
    assert payload["current_absolute_index"] == 16

    review_page = client.get("/ui/review", params={"status": "pending", "page": 2, "page_size": 15})
    assert review_page.status_code == 200
    assert 'value="16 / 30"' in review_page.text
    assert "item 16 / 30" in review_page.text
    assert "Selected 16 / 30" in review_page.text

    detail = client.get("/ui/events/run_page_00_e1?camera_id=cam_01&status=pending&page=2&page_size=15")
    assert detail.status_code == 200
    assert "/ui/review?camera_id=cam_01&amp;status=pending&amp;event_uid=run_page_00_e1&amp;page=2&amp;page_size=15" in detail.text


def test_review_queue_only_enriches_visible_page_and_adjacent_items(
    monkeypatch,
    tmp_path,
) -> None:
    for index in range(40):
        _write_run(
            tmp_path,
            day="2026-03-11",
            run_uid=f"run_bounded_{index:02d}",
            started_at_utc=f"2026-03-11T10:{index:02d}:00Z",
            occurred_at_utc=f"2026-03-11T10:{index:02d}:30Z",
        )

    app = create_app(
        spool_dir=tmp_path,
        review_db_path=tmp_path / "reviews.sqlite3",
    )
    runtime = app.state.runtime
    original_attach = runtime._attach_review_data
    enriched_batch_sizes = []

    def record_attach(items, *, review_map=None):
        enriched_batch_sizes.append(len(items))
        return original_attach(items, review_map=review_map)

    monkeypatch.setattr(runtime, "_attach_review_data", record_attach)
    monkeypatch.setattr(
        runtime.review_store,
        "get_reviews",
        lambda event_uids: (_ for _ in ()).throw(
            AssertionError("backlog-sized IN query used")
        ),
    )

    payload = runtime.review_queue_payload(
        status_filter="pending",
        page=2,
        page_size=15,
    )

    assert payload["queue_total"] == 40
    assert payload["page_item_count"] == 15
    assert enriched_batch_sizes == [16]


def test_review_queue_preserves_filtered_status_counts(tmp_path) -> None:
    for run_uid, minute, camera_id in (
        ("run_count_yes", "00", "cam_01"),
        ("run_count_no", "01", "cam_02"),
        ("run_count_pending", "02", "cam_01"),
    ):
        _write_run(
            tmp_path,
            day="2026-03-11",
            run_uid=run_uid,
            started_at_utc=f"2026-03-11T10:{minute}:00Z",
            occurred_at_utc=f"2026-03-11T10:{minute}:30Z",
            camera_id=camera_id,
        )

    app = create_app(
        spool_dir=tmp_path,
        review_db_path=tmp_path / "reviews.sqlite3",
    )
    store = app.state.runtime.review_store
    store.save_review(
        event_uid="run_count_yes_e1",
        run_uid="run_count_yes",
        site_id="site_a",
        camera_id="cam_01",
        decision="qualified_yes",
        reviewed_class=None,
        notes="",
        now_utc="2026-03-11T10:03:00Z",
    )
    store.save_review(
        event_uid="run_count_no_e1",
        run_uid="run_count_no",
        site_id="site_a",
        camera_id="cam_02",
        decision="qualified_no",
        reviewed_class=None,
        notes="",
        now_utc="2026-03-11T10:04:00Z",
    )

    payload = app.state.runtime.review_queue_payload(
        camera_id="cam_01",
        status_filter="all",
    )

    assert payload["queue_total"] == 2
    assert payload["cameras"] == ["cam_01", "cam_02"]
    assert payload["counts"] == {
        "pending": 1,
        "qualified_yes": 1,
        "qualified_no": 0,
        "reviewed_total": 1,
    }
    assert [item["event_uid"] for item in payload["items"]] == [
        "run_count_yes_e1",
        "run_count_pending_e1",
    ]


def test_review_queue_excel_export_uses_date_slice_and_builds_weight_summary(tmp_path) -> None:
    _write_run(
        tmp_path,
        day="2026-03-10",
        run_uid="run_before",
        started_at_utc="2026-03-10T10:00:00Z",
        occurred_at_utc="2026-03-10T10:05:00Z",
    )
    _write_run(
        tmp_path,
        day="2026-03-11",
        run_uid="run_accepted",
        started_at_utc="2026-03-11T10:00:00Z",
        occurred_at_utc="2026-03-11T10:05:00Z",
        direction="B_TO_A",
        class_name="pickup",
    )
    _write_run(
        tmp_path,
        day="2026-03-11",
        run_uid="run_pending",
        started_at_utc="2026-03-11T10:10:00Z",
        occurred_at_utc="2026-03-11T10:15:00Z",
        class_name="tronton",
    )
    _write_run(
        tmp_path,
        day="2026-03-11",
        run_uid="run_rejected",
        started_at_utc="2026-03-11T10:20:00Z",
        occurred_at_utc="2026-03-11T10:25:00Z",
        class_name="trailer",
    )
    app = create_app(spool_dir=tmp_path, review_db_path=tmp_path / "reviews.sqlite3")
    store = app.state.runtime.review_store
    store.save_review(
        event_uid="run_accepted_e1",
        run_uid="run_accepted",
        site_id="site_a",
        camera_id="cam_01",
        decision="qualified_yes",
        reviewed_class=None,
        notes="",
        now_utc="2026-03-11T10:30:00Z",
    )
    store.save_review(
        event_uid="run_rejected_e1",
        run_uid="run_rejected",
        site_id="site_a",
        camera_id="cam_01",
        decision="qualified_no",
        reviewed_class=None,
        notes="",
        now_utc="2026-03-11T10:31:00Z",
    )
    client = TestClient(app)

    review_page = client.get("/ui/review", params={"date_from": "2026-03-11", "date_to": "2026-03-11"})
    assert review_page.status_code == 200
    assert (
        '/ui/review/export.xlsx?date_from=2026-03-11&amp;date_to=2026-03-11'
        in review_page.text
    )
    assert review_page.text.index("Download Excel") < review_page.text.index("Buka halaman detail")

    response = client.get(
        "/ui/review/export.xlsx",
        params={"date_from": "2026-03-11", "date_to": "2026-03-11", "status": "qualified_no"},
    )
    assert response.status_code == 200
    assert response.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert response.headers["content-disposition"] == (
        'attachment; filename="antrian-review_2026-03-11.xlsx"'
    )

    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == ["Antrian Review", "Summary"]

    detail = workbook["Antrian Review"]
    rows = list(detail.iter_rows(values_only=True))
    assert rows == [
        (
            "Waktu (WIB)",
            "Camera",
            "Arah",
            "Tipe Kendaraan",
            "Estimasi Berat Muatan(T)",
            "Akurasi (%)",
            "Review",
        ),
        (
            "2026-03-11 17:05 WIB",
            "cam_01",
            "Finished Goods",
            "pickup",
            1.5,
            91,
            "Diterima",
        ),
    ]
    assert detail.auto_filter.ref == "A1:G2"

    summary = workbook["Summary"]
    assert summary.freeze_panes == "A3"
    assert summary.auto_filter.ref == "A2:D9"
    assert [str(cell_range) for cell_range in summary.merged_cells.ranges] == ["B1:D1"]
    assert [summary.cell(row=row_number, column=1).value for row_number in range(3, 10)] == [
        "pickup",
        "double engkel",
        "engkel",
        "tronton",
        "trailer",
        "container 20ft",
        "container 40ft",
    ]
    assert summary["B3"].value == (
        "=SUMIFS("
        "'Antrian Review'!$E:$E,"
        "'Antrian Review'!$D:$D,$A3,"
        "'Antrian Review'!$C:$C,B$2"
        ")"
    )
    assert summary["C3"].value == (
        "=SUMIFS("
        "'Antrian Review'!$E:$E,"
        "'Antrian Review'!$D:$D,$A3,"
        "'Antrian Review'!$C:$C,C$2"
        ")"
    )
    assert summary["D3"].value == "=SUM(B3:C3)"
    assert (
        summary["A10"].value,
        summary["B10"].value,
        summary["C10"].value,
        summary["D10"].value,
    ) == (
        "Grand Total",
        "=SUM(B3:B9)",
        "=SUM(C3:C9)",
        "=SUM(B10:C10)",
    )
    for column in "ABCD":
        header_cell = summary[f"{column}2"]
        total_cell = summary[f"{column}10"]
        assert total_cell.font.bold is True
        assert total_cell.fill.fill_type == header_cell.fill.fill_type
        assert total_cell.fill.fgColor.rgb == header_cell.fill.fgColor.rgb


def test_event_detail_includes_cross_page_review_navigation(tmp_path) -> None:
    for index in range(30):
        _write_run(
            tmp_path,
            day="2026-03-11",
            run_uid=f"run_nav_{index:02d}",
            started_at_utc=f"2026-03-11T10:{index:02d}:00Z",
            occurred_at_utc=f"2026-03-11T10:{index:02d}:30Z",
        )

    client = TestClient(create_app(spool_dir=tmp_path, review_db_path=tmp_path / "reviews.sqlite3"))

    page_one = client.get("/review/queue", params={"status": "pending", "page": 1, "page_size": 15})
    assert page_one.status_code == 200
    boundary_event_uid = page_one.json()["items"][-1]["event_uid"]

    focused_queue = client.get(
        "/review/queue",
        params={"status": "pending", "event_uid": boundary_event_uid, "page": 1, "page_size": 15},
    )
    assert focused_queue.status_code == 200
    queue_payload = focused_queue.json()
    assert queue_payload["current"]["event_uid"] == boundary_event_uid
    assert queue_payload["current_absolute_index"] == 15
    assert queue_payload["next_item"]["detail_url"].endswith("&page=2&page_size=15")

    detail = client.get(f"/ui/events/{boundary_event_uid}?status=pending&page=1&page_size=15")
    assert detail.status_code == 200
    assert "item 15 / 30" in detail.text
    assert "lanjut otomatis" in detail.text
    assert "&amp;page=2&amp;page_size=15" in detail.text
    assert "data-next-detail-url=" in detail.text


def test_favicon_redirects_to_static_asset(tmp_path) -> None:
    client = TestClient(create_app(spool_dir=tmp_path))

    response = client.get("/favicon.ico", follow_redirects=False)
    assert response.status_code == 307
    assert response.headers["location"] == "/ui-static/favicon.svg"



def test_ui_login_gates_pages_and_sets_cookie(tmp_path) -> None:
    _write_run(
        tmp_path,
        day="2026-03-11",
        run_uid="run_ui",
        started_at_utc="2026-03-11T10:00:00Z",
        occurred_at_utc="2026-03-11T10:05:00Z",
    )
    client = TestClient(
        create_app(
            spool_dir=tmp_path,
            review_db_path=tmp_path / "reviews.sqlite3",
            ui_auth_cfg=UiAuthConfig(username="admin", password="secret"),
        )
    )

    login_page = client.get("/ui/login")
    assert login_page.status_code == 200
    assert "Sign In" in login_page.text

    gated = client.get("/ui/dashboard", follow_redirects=False)
    assert gated.status_code == 307
    assert gated.headers["location"].startswith("/ui/login")

    failed_login = client.post("/api/auth/login", json={"username": "admin", "password": "wrong"})
    assert failed_login.status_code == 401

    good_login = client.post("/api/auth/login", json={"username": "admin", "password": "secret"})
    assert good_login.status_code == 200
    assert "edge_ui_session" in good_login.headers.get("set-cookie", "")

    dashboard = client.get("/ui/dashboard")
    assert dashboard.status_code == 200
    assert "Traffic Monitoring Dashboard" in dashboard.text

    review_resp = client.post(
        "/events/run_ui_e1/review",
        json={
            "decision": "qualified_no",
            "reject_reason": "not_relevant",
            "notes": "not target",
        },
    )
    assert review_resp.status_code == 200
    assert review_resp.json()["review"]["decision"] == "qualified_no"


def test_ui_login_rate_limits_repeated_failed_json_attempts(tmp_path) -> None:
    client = TestClient(
        create_app(
            spool_dir=tmp_path,
            ui_auth_cfg=UiAuthConfig(
                username="admin",
                password="secret",
                login_rate_limit_max_failures=3,
                login_rate_limit_window_s=300,
                login_rate_limit_lockout_s=60,
            ),
        )
    )

    for _ in range(2):
        failed_login = client.post("/api/auth/login", json={"username": "admin", "password": "wrong"})
        assert failed_login.status_code == 401
        assert failed_login.json()["detail"] == "invalid_credentials"

    blocked_login = client.post("/api/auth/login", json={"username": "admin", "password": "wrong"})
    assert blocked_login.status_code == 429
    assert blocked_login.json()["detail"] == "too_many_attempts"
    assert blocked_login.headers["retry-after"] == "60"

    blocked_good_login = client.post("/api/auth/login", json={"username": "admin", "password": "secret"})
    assert blocked_good_login.status_code == 429
    assert blocked_good_login.json()["detail"] == "too_many_attempts"


def test_ui_login_rate_limit_resets_after_success(tmp_path) -> None:
    client = TestClient(
        create_app(
            spool_dir=tmp_path,
            ui_auth_cfg=UiAuthConfig(
                username="admin",
                password="secret",
                login_rate_limit_max_failures=3,
                login_rate_limit_window_s=300,
                login_rate_limit_lockout_s=60,
            ),
        )
    )

    for _ in range(2):
        failed_login = client.post("/api/auth/login", json={"username": "admin", "password": "wrong"})
        assert failed_login.status_code == 401

    good_login = client.post("/api/auth/login", json={"username": "admin", "password": "secret"})
    assert good_login.status_code == 200

    for _ in range(2):
        failed_login = client.post("/api/auth/login", json={"username": "admin", "password": "wrong"})
        assert failed_login.status_code == 401

    still_allowed = client.post("/api/auth/login", json={"username": "admin", "password": "secret"})
    assert still_allowed.status_code == 200


def test_ui_login_form_redirects_with_error_and_success(tmp_path) -> None:
    _write_run(
        tmp_path,
        day="2026-03-11",
        run_uid="run_ui_form",
        started_at_utc="2026-03-11T10:00:00Z",
        occurred_at_utc="2026-03-11T10:05:00Z",
    )
    client = TestClient(
        create_app(
            spool_dir=tmp_path,
            review_db_path=tmp_path / "reviews.sqlite3",
            ui_auth_cfg=UiAuthConfig(username="admin", password="secret"),
        )
    )

    failed_login = client.post(
        "/ui/login",
        data={"username": "admin", "password": "wrong", "next": "/ui/review?status=pending"},
        follow_redirects=False,
    )
    assert failed_login.status_code == 303
    assert failed_login.headers["location"] == (
        "/ui/login?next=%2Fui%2Freview%3Fstatus%3Dpending&error=invalid_credentials&username=admin"
    )

    failed_page = client.get(failed_login.headers["location"])
    assert failed_page.status_code == 200
    assert "Login failed. Check your credentials and try again." in failed_page.text
    assert 'value="admin"' in failed_page.text

    good_login = client.post(
        "/ui/login",
        data={"username": "admin", "password": "secret", "next": "/ui/review?status=pending"},
        follow_redirects=False,
    )
    assert good_login.status_code == 303
    assert good_login.headers["location"] == "/ui/review?status=pending"
    assert "edge_ui_session" in good_login.headers.get("set-cookie", "")

    review_page = client.get("/ui/review?status=pending")
    assert review_page.status_code == 200
    assert "Antrian Review" in review_page.text


def test_ui_login_form_redirects_when_rate_limited(tmp_path) -> None:
    client = TestClient(
        create_app(
            spool_dir=tmp_path,
            ui_auth_cfg=UiAuthConfig(
                username="admin",
                password="secret",
                login_rate_limit_max_failures=2,
                login_rate_limit_window_s=300,
                login_rate_limit_lockout_s=60,
            ),
        )
    )

    first_failed_login = client.post(
        "/ui/login",
        data={"username": "admin", "password": "wrong", "next": "/ui/dashboard"},
        follow_redirects=False,
    )
    assert first_failed_login.status_code == 303
    assert first_failed_login.headers["location"] == "/ui/login?next=%2Fui%2Fdashboard&error=invalid_credentials&username=admin"

    blocked_login = client.post(
        "/ui/login",
        data={"username": "admin", "password": "wrong", "next": "/ui/dashboard"},
        follow_redirects=False,
    )
    assert blocked_login.status_code == 303
    assert blocked_login.headers["location"] == "/ui/login?next=%2Fui%2Fdashboard&error=too_many_attempts&username=admin"

    blocked_page = client.get(blocked_login.headers["location"])
    assert blocked_page.status_code == 200
    assert "Too many login attempts. Wait a bit before trying again." in blocked_page.text


def test_get_event_uses_catalog_without_scanning_jsonl(monkeypatch, tmp_path) -> None:
    _write_run(
        tmp_path,
        day="2026-03-11",
        run_uid="run_catalog_lookup",
        started_at_utc="2026-03-11T10:00:00Z",
        occurred_at_utc="2026-03-11T10:05:00Z",
    )
    app = create_app(
        spool_dir=tmp_path,
        review_db_path=tmp_path / "reviews.sqlite3",
    )

    def fail_if_jsonl_is_scanned(path: Path):
        raise AssertionError(f"legacy JSONL scan used: {path}")

    monkeypatch.setattr(
        api_module,
        "_iter_jsonl_records",
        fail_if_jsonl_is_scanned,
        raising=False,
    )

    event = app.state.runtime.get_event("run_catalog_lookup_e1")

    assert event is not None
    assert event["event_uid"] == "run_catalog_lookup_e1"
    assert event["run"]["run_uid"] == "run_catalog_lookup"
    assert event["review_status"] == "pending"
    assert event["timeline"]

def test_iter_all_events_uses_ordered_catalog_without_scanning_jsonl(
    monkeypatch,
    tmp_path,
) -> None:
    _write_run(
        tmp_path,
        day="2026-03-11",
        run_uid="run_new",
        started_at_utc="2026-03-11T10:10:00Z",
        occurred_at_utc="2026-03-11T10:15:00Z",
    )
    _write_run(
        tmp_path,
        day="2026-03-11",
        run_uid="run_old",
        started_at_utc="2026-03-11T10:00:00Z",
        occurred_at_utc="2026-03-11T10:05:00Z",
    )
    app = create_app(
        spool_dir=tmp_path,
        review_db_path=tmp_path / "reviews.sqlite3",
    )

    def fail_if_jsonl_is_scanned(path: Path):
        raise AssertionError(f"legacy JSONL scan used: {path}")

    monkeypatch.setattr(
        api_module,
        "_iter_jsonl_records",
        fail_if_jsonl_is_scanned,
        raising=False,
    )

    events = list(app.state.runtime._iter_all_events())

    assert [event["event_uid"] for event in events] == [
        "run_old_e1",
        "run_new_e1",
    ]


def test_list_reviewable_classes_uses_catalog_without_scanning_jsonl(
    monkeypatch,
    tmp_path,
) -> None:
    _write_run(
        tmp_path,
        day="2026-03-11",
        run_uid="run_catalog_classes",
        started_at_utc="2026-03-11T10:00:00Z",
        occurred_at_utc="2026-03-11T10:05:00Z",
        class_name="observed truck",
        run_class_names={0: "taxonomy pickup"},
    )
    app = create_app(
        spool_dir=tmp_path,
        review_db_path=tmp_path / "reviews.sqlite3",
    )
    app.state.runtime.review_store.save_review(
        event_uid="previously_reviewed_event",
        run_uid=None,
        site_id=None,
        camera_id=None,
        decision="qualified_yes",
        reviewed_class="reviewed trailer",
        notes="",
        now_utc="2026-03-11T10:06:00Z",
    )

    def fail_if_jsonl_is_scanned(path: Path):
        raise AssertionError(f"legacy JSONL scan used: {path}")

    monkeypatch.setattr(
        api_module,
        "_iter_jsonl_records",
        fail_if_jsonl_is_scanned,
        raising=False,
    )

    class_names = app.state.runtime.list_reviewable_classes()

    assert class_names == [
        "observed truck",
        "reviewed trailer",
        "taxonomy pickup",
    ]
